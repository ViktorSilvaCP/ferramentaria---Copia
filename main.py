from flask import Flask, render_template, request, redirect, url_for, flash,send_file, session, jsonify; import sqlite3; import pandas as pd;from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, login_required, login_user, logout_user, current_user
from werkzeug.utils import secure_filename
from math import ceil
from flask_sqlalchemy import SQLAlchemy
from werkzeug.exceptions import Unauthorized
from datetime import datetime, timedelta  
import os
import sys
import logging
from utils.datetime_helper import get_current_datetime, format_datetime, parse_datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, timedelta
import traceback
from collections import defaultdict
import numpy as np
import logging.handlers
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from apscheduler.schedulers.background import BackgroundScheduler
import config
from jinja2 import Template
from apscheduler.jobstores.base import JobLookupError
from apscheduler.triggers.cron import CronTrigger
import pytz
import re
from io import BytesIO
import uuid
from email.mime.image import MIMEImage
# IMPORTAÇÕES DO PYTHON
import time
import threading
from utils.ferramentas_importer import consumir_ferramentas, importar_ferramentas_para_db
log_formatter = logging.Formatter('%(asctime)s [%(levelname)s] [%(filename)s:%(lineno)d] %(message)s')
log_file = 'app.log'
file_handler = logging.handlers.RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5)
file_handler.setFormatter(log_formatter)
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
logger.addHandler(file_handler)
logger.addHandler(console_handler)
base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
template_path = os.path.join(base_path, "templates")
static_path = os.path.join(base_path, "static")
relatorio_path = os.path.join(base_path, "relatorio")
app = Flask(__name__, template_folder=template_path, static_folder=static_path)
app.jinja_env.add_extension('jinja2.ext.do')
app.secret_key = 'a7da6230bce08f317c2a0de0d383740c7f21b168a7da6230bce08f317c2a0de0d383740c7f21b168'
app.permanent_session_lifetime = timedelta(minutes=15)
app.config['SESSION_PERMANENT'] = True
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///hotspots.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'pdf', 'mp4', 'avi', 'mov', 'mkv', 'jfif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
login_manager = LoginManager()
login_manager.init_app(app)



class Hotspot(db.Model):
    __tablename__ = 'hotspots'  
    
    id = db.Column(db.Integer, primary_key=True)
    top = db.Column(db.String(50), nullable=False)
    left = db.Column(db.String(50), nullable=False)
    posicao = db.Column(db.Integer, nullable=False) 

    def __repr__(self):
        return f"<Hotspot posição {self.posicao}>"

class Afiação(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    posicao = db.Column(db.Integer, nullable=False)
    ferramenta = db.Column(db.String(50), nullable=False)
    lado = db.Column(db.String(50), nullable=False)
    altura = db.Column(db.Float, nullable=False)
    folga = db.Column(db.Float, nullable=False)
    spacer = db.Column(db.String(50), nullable=True)
    data_troca = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    dias_produzidos = db.Column(db.Integer, nullable=False)
    ferramenteiro = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"<Afiação {self.posicao} - {self.ferramenta}>"
class HistoricoTroca(db.Model):
    __tablename__ = 'historico_troca'  
    id = db.Column(db.Integer, primary_key=True)
    posicao = db.Column(db.Integer, nullable=False)
    codigo = db.Column(db.String(50), nullable=False)
    operador = db.Column(db.String(100), nullable=False)
    data = db.Column(db.DateTime(timezone=True), default=get_current_datetime)  
    vida_util = db.Column(db.Integer)
    producao_atual = db.Column(db.Integer, default=0)

    def __repr__(self):
        return f"<HistoricoTroca {self.codigo} na posição {self.posicao}>"


class Ferramenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    tipo = db.Column(db.String(50), nullable=False)
    descricao = db.Column(db.String(200), nullable=True)
    status = db.Column(db.String(20), default='disponivel')  
    posicao = db.Column(db.Integer, nullable=True)  
    ultima_atualizacao = db.Column(db.DateTime(timezone=True), default=get_current_datetime)  
    dimensao_metrica = db.Column(db.String(50), nullable=True)
    dimensao_polegada = db.Column(db.String(50), nullable=True)
    sufixo = db.Column(db.String(20), nullable=True)
    

class Faca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    posicao = db.Column(db.Integer, nullable=False)
    ferramenta = db.Column(db.String(100), nullable=False)
    lado = db.Column(db.String(1), nullable=False)  
    altura = db.Column(db.Float, nullable=False)
    folga = db.Column(db.Float, nullable=True)
    spacer = db.Column(db.String(50), nullable=True)
    data_troca = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    dias_produzidos = db.Column(db.Integer, default=0)
    utilizador = db.Column(db.String(100), nullable=False)


class HistoricoFacas(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    posicao = db.Column(db.Integer, nullable=False)
    ferramenta_anterior = db.Column(db.String(100))
    ferramenta_nova = db.Column(db.String(100), nullable=False)
    lado = db.Column(db.String(1), nullable=False)
    altura = db.Column(db.Float, nullable=False)
    folga = db.Column(db.Float, nullable=True)
    spacer = db.Column(db.String(50), nullable=True)
    data_troca = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    utilizador = db.Column(db.String(100), nullable=False)


class ManutencaoFerramenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ferramenta_id = db.Column(db.Integer, db.ForeignKey('ferramenta.id'), nullable=False)
    data_entrada = db.Column(db.DateTime(timezone=True), default=get_current_datetime)  
    data_saida = db.Column(db.DateTime(timezone=True), nullable=True)
    motivo_entrada = db.Column(db.String(200), nullable=False)
    descricao_manutencao = db.Column(db.String(500), nullable=True)
    status = db.Column(db.String(20), default='pendente')  
    operador_entrada = db.Column(db.String(100), nullable=False)
    operador_saida = db.Column(db.String(100), nullable=True)
    motivo_descarte = db.Column(db.String(500), nullable=True)  
    
    ferramenta = db.relationship('Ferramenta', backref=db.backref('manutencoes', lazy=True))
class DescarteFerramenta(db.Model):
    __tablename__ = 'descarte_ferramenta'
    id = db.Column(db.Integer, primary_key=True)
    ferramenta_id = db.Column(db.Integer, db.ForeignKey('ferramenta.id'))
    codigo = db.Column(db.String(50), nullable=False)
    operador = db.Column(db.String(100), nullable=False)
    motivo = db.Column(db.Text, nullable=False)
    data_descarte = db.Column(db.DateTime(timezone=True), nullable=False, default=get_current_datetime)
    
    def __repr__(self):
        return f'<DescarteFerramenta {self.codigo}>'


class Admin:
    def __init__(self, id, matricula, senha, nome=None, area=None):
        self.id = id
        self.matricula = matricula
        self.senha = senha
        self.nome = nome 
        self.area = area
    def get_id(self):
        return str(self.id)

    
    @property
    def is_active(self):
        return True  

    @property
    def is_authenticated(self):
        return True  

    @property
    def is_anonymous(self):
        return False  
        
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.errorhandler(Unauthorized)
def handle_unauthorized(error):
    return redirect(url_for('login'))

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    c.execute("SELECT id, matricula, senha, nome, area FROM admin WHERE id = ?", (user_id,))
    user_data = c.fetchone()
    conn.close()
    if user_data:
        return Admin(user_data[0], user_data[1], user_data[2], user_data[3], user_data[4])
    return None


def init_db():
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()

    
    c.execute('''CREATE TABLE IF NOT EXISTS admin (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL UNIQUE,
        matricula TEXT NOT NULL UNIQUE,
        senha TEXT,
        area TEXT CHECK(area IN ('tampas', 'latas', 'supervisor')) NOT NULL DEFAULT 'latas'
    )''')

    
    c.execute('''CREATE TABLE IF NOT EXISTS historico (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT,
        nome TEXT,
        tipo_acao TEXT,
        equipamento TEXT,
        solicitante TEXT,
        codigo_falha TEXT,
        causa_encontrada TEXT,
        comentario TEXT,
        horario_inicio TEXT,
        horario_termino TEXT,
        foto BLOB,
        eficiencia REAL,
        area TEXT CHECK(area IN ('tampas', 'latas')) NOT NULL DEFAULT 'latas'
    )''')

    
    c.execute('''CREATE TABLE IF NOT EXISTS historico_backup (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_original INTEGER,
        data TEXT,
        nome TEXT,
        tipo_acao TEXT,
        equipamento TEXT,
        solicitante TEXT,
        codigo_falha TEXT,
        causa_encontrada TEXT,
        trabalho_executado TEXT,
        comentario TEXT,
        horario_inicio TEXT,
        horario_termino TEXT,
        foto BLOB,
        eficiencia REAL,
        editado_por TEXT,
        data_edicao TEXT
    )''')

    
    c.execute('''CREATE TABLE IF NOT EXISTS edit_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        registro_id INTEGER,
        data_edicao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        editor_id INTEGER,
        dados_anteriores TEXT,
        FOREIGN KEY(registro_id) REFERENCES historico(id),
        FOREIGN KEY(editor_id) REFERENCES admin(id)
    )''')

    conn.commit()
    conn.close()

with app.app_context():
    db.create_all()

AVAILABLE_TOOL_TYPES = {
    'DCP': 'DIE CENTER PISTON',
    'BDP': 'BLANK-DRAW PUNCH',
    'CTE': 'CUT EDGE',
    'DCA': 'DIE CENTER ASSEMBLY',
    'DCR': 'DIE CORE RING',
    'INP': 'INNER PRESSURE',
    'LWP': 'LOWER PISTON',
    'LWR': 'LOWER RETAINER',
    'PNP': 'PANEL PUNCH',
    'PPP': 'PANEL PUNCH PISTON',
    'UPP': 'UPPER PISTON',
    'UPR': 'UPPER RETAINER'
}


@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        matricula = request.form['matricula']
        senha = request.form['senha']
        print(f"Dados enviados: Matrícula: {matricula}, Senha: {senha}")
        conn = sqlite3.connect('relatorio_diario.db')
        c = conn.cursor()
        c.execute("SELECT id, matricula, senha FROM admin WHERE matricula = ?", (matricula,))
        usuario = c.fetchone()
        conn.close()
        if usuario:
            print(f"Usuário encontrado: {usuario}")
            if check_password_hash(usuario[2], senha):  
                print("Senha validada com sucesso.")
                
                user = Admin(usuario[0], usuario[1], usuario[2])
                login_user(user)
                session.permanent = True
                
                if senha == 'canpack.2025':  
                    flash('Por favor, redefina sua senha.', 'info')
                    return redirect(url_for('definir_senha', matricula=matricula))
                return redirect(url_for('index'))
            else:
                print("Senha inválida.")
                flash('Credenciais inválidas.', 'danger')
        else:
            print("Usuário não encontrado no banco de dados.")
            flash('Credenciais inválidas.', 'danger')
    return render_template('login.html')


@app.route('/redefinir_senha', methods=['POST'])
def redefinir_senha():
    
    matricula = request.form.get('matricula')

    if not matricula:
        flash('Matrícula é obrigatória para redefinir a senha.', 'danger')
        return redirect(url_for('index'))

    
    nova_senha = 'canpack.2025'
    senha_hash = generate_password_hash(nova_senha)

    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    c.execute("UPDATE admin SET senha = ? WHERE matricula = ?", (senha_hash, matricula))
    conn.commit()
    conn.close()

    flash(f'A senha do usuário com matrícula {matricula} foi redefinida.', 'success')
    return redirect(url_for('adicionar_matricula'))


@app.route('/index')
@login_required
def index():
    session.permanent = True
    
    
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    c.execute("SELECT area FROM admin WHERE matricula = ?", (current_user.matricula,))
    user_area = c.fetchone()
    conn.close()

    
    if user_area and user_area[0]:
        
        roles = [role.strip() for role in user_area[0].split(',')] if ',' in user_area[0] else [user_area[0]]
    else:
        roles = []

    
    current_user.roles = roles
    
    return render_template('index.html')

@app.route('/production')
@login_required
def production():  
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    session.permanent = True
    return render_template('production.html')

@app.route('/afiacoes')
@login_required
def afiacoes_view():  
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    session.permanent = True
    return render_template('afiacoes.html')

@app.route('/ferramentas')
@login_required
def ferramentas():  
    """
    Rota para exibir página de gerenciamento de ferramentas disponíveis.
    """
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    session.permanent = True
    logging.info(f"[FERRAMENTAS] Usuário {current_user.nome} acessou página de ferramentas")
    return render_template('ferramentas.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu com sucesso.', 'success')
    return redirect(url_for('login'))


@app.route('/definir_senha/<matricula>', methods=['GET', 'POST'])
def definir_senha(matricula):
    if request.method == 'POST':
        nova_senha = request.form['senha']
        senha_hash = generate_password_hash(nova_senha)

        conn = sqlite3.connect('relatorio_diario.db')
        c = conn.cursor()
        c.execute("UPDATE admin SET senha = ? WHERE matricula = ?", (senha_hash, matricula))
        conn.commit()
        conn.close()

        flash('Senha redefinida com sucesso!', 'success')
        return redirect(url_for('login'))
    return render_template('definir_senha.html', matricula=matricula)

@app.route('/editar_relatorio/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_relatorio(id):
    logging.debug(f"Rota '/editar_relatorio/{id}' acessada com método {request.method}")
    user_id = current_user.get_id()
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()

    if request.method == 'POST':
        try:
            
            foto = request.files.get('foto')
            foto_path = None
            if foto and allowed_file(foto.filename):
                foto_nome = secure_filename(foto.filename)
                foto_path = os.path.join(app.config['UPLOAD_FOLDER'], foto_nome)
                foto.save(foto_path)

            
            form_data = request.form.to_dict()
            
            
            form_data.pop('area', None)  

            
            if 'data' in form_data:
                try:
                    data_parts = form_data['data'].split('-')
                    if len(data_parts) == 3:
                        form_data['data'] = f"{data_parts[0]}-{data_parts[1]}-{data_parts[2]}"
                except Exception as e:
                    logging.error(f"Erro ao converter data: {e}")

            
            c.execute('''SELECT id, data, nome, tipo_acao, equipamento, solicitante, 
                        codigo_falha, causa_encontrada, trabalho_executado, comentario, 
                        horario_inicio, horario_termino, foto, eficiencia 
                        FROM historico WHERE id = ?''', (id,))
            registro_original = c.fetchone()

            if not registro_original:
                return jsonify(success=False, message='Relatório não encontrado.')

            
            logging.debug(f"Dados originais: {registro_original}")
            logging.debug(f"Dados do formulário: {form_data}")

            
            c.execute('''
                INSERT INTO historico_backup (
                    id_original, data, nome, tipo_acao, equipamento, solicitante,
                    codigo_falha, causa_encontrada, trabalho_executado, comentario,
                    horario_inicio, horario_termino, foto, eficiencia,
                    editado_por, data_edicao
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ''', (
                registro_original[0],  
                registro_original[1],  
                registro_original[2],  
                registro_original[3],  
                registro_original[4],  
                registro_original[5],  
                registro_original[6],  
                registro_original[7],  
                registro_original[8],  
                registro_original[9],  
                registro_original[10], 
                registro_original[11], 
                registro_original[12], 
                registro_original[13], 
                current_user.nome      
            ))

            
            campos = []
            valores = []
            for campo, valor in form_data.items():
                if valor is not None and campo != "id":
                    campos.append(f"{campo} = ?")
                    valores.append(valor)

            
            if foto_path:
                campos.append("foto = ?")
                valores.append(foto_path)

            valores.append(id)
            query = f"UPDATE historico SET {', '.join(campos)} WHERE id = ?"
            c.execute(query, tuple(valores))
            
            
            conn.commit()

            return jsonify(
                success=True, 
                message="Relatório atualizado com sucesso e backup criado!", 
                nome_editor=current_user.nome
            )

        except Exception as e:
            conn.rollback()
            logging.error(f"Erro ao atualizar relatório: {str(e)}")
            return jsonify(
                success=False, 
                message=f"Erro ao atualizar o relatório: {str(e)}"
            )
        finally:
            conn.close()

    return jsonify(success=False, message='Método não permitido')

from datetime import datetime

@app.route('/historico_acoes', methods=['GET', 'POST'])
@login_required
def historico_acoes():
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    search_filter = request.args.get('search')  
    data_filter = request.args.get('data')  
    nome_filter = request.args.get('nome')
    horario_inicio_filter = request.args.get('horario_inicio')
    tipo_acao_filter = request.args.get('tipo_acao')
    equipamento_filter = request.args.get('equipamento')
    solicitante_filter = request.args.get('solicitante')
    codigo_falha_filter = request.args.get('codigo_falha')
    page = request.args.get('page', 1, type=int)
    area_filter = request.args.get('area')
    if data_filter:
        try:
            data_filter = datetime.strptime(data_filter, '%Y-%m-%d').strftime('%Y-%m-%d')  
        except ValueError:
            data_filter = None  
    query = """
        SELECT
            h.id,
            h.data,
            h.nome,
            h.horario_inicio,
            h.tipo_acao,
            h.equipamento,
            h.solicitante,
            h.codigo_falha,
            h.causa_encontrada,
            h.trabalho_executado,
            h.comentario,
            h.foto,
            h.area
        FROM historico h
        WHERE 1=1
    """

    
    params = []
    if search_filter:
        query += " AND (id = ? OR nome LIKE ?)"
        params.append(search_filter)
        params.append(f'%{search_filter}%')
    if data_filter:  
        query += " AND DATE(data) = ?"  
        params.append(data_filter)
    if nome_filter:
        query += " AND nome LIKE ?"
        params.append(f'%{nome_filter}%')
    if horario_inicio_filter:
        query += " AND horario_inicio = ?"
        params.append(horario_inicio_filter)
    if tipo_acao_filter:
        query += " AND tipo_acao LIKE ?"
        params.append(f'%{tipo_acao_filter}%')
    if equipamento_filter:
        query += " AND equipamento LIKE ?"
        params.append(f'%{equipamento_filter}%')
    if solicitante_filter:
        query += " AND solicitante LIKE ?"
        params.append(f'%{solicitante_filter}%')
    if codigo_falha_filter:
        query += " AND codigo_falha LIKE ?"
        params.append(f'%{codigo_falha_filter}%')

    
    if current_user.area == 'supervisor':
        if area_filter:  
            query += " AND h.area = ?"
            params.append(area_filter)
    else:  
        query += " AND h.area = ?"
        params.append(current_user.area)

    
    count_query = f"SELECT COUNT(*) FROM ({query}) as count_table"
    
    
    logging.debug(f"Query: {count_query}")
    logging.debug(f"Params: {params}")
    logging.debug(f"User area: {current_user.area}")

    try:
        c.execute(count_query, tuple(params))
        result = c.fetchone()
        total_records = result[0] if result else 0

        
        records_per_page = 40
        total_pages = ceil(total_records / records_per_page)

        
        logging.debug(f"Total records: {total_records}")
        logging.debug(f"Records per page: {records_per_page}")
        logging.debug(f"Total pages: {total_pages}")

        
        offset = (page - 1) * records_per_page
        query += " ORDER BY h.data DESC, h.horario_inicio DESC"
        query += f" LIMIT {records_per_page} OFFSET {offset}"

        
        c.execute(query, tuple(params))
        registros = c.fetchall()

        
        registros_corrigidos = [
            [
                
                datetime.strptime(col, '%Y-%m-%d').strftime('%d-%m-%Y') if i == 1 and col else col
                for i, col in enumerate(registro[:-1])  
            ] + [registro[-1]]  
            for registro in registros
        ]

        
        for registro in registros_corrigidos:
            if registro[11] is None:
                registro[11] = "N/A"  
        conn.close()

    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        flash('Erro ao carregar os registros.', 'danger')
        registros_corrigidos = []
        total_pages = 1
        page = 1

    finally:
        conn.close()

    
    return render_template(
        'historico_acoes.html', 
        registros=registros_corrigidos, 
        page=page, 
        total_pages=total_pages,
        search_filter=search_filter,  
        data_filter=data_filter,
        nome_filter=nome_filter,
        horario_inicio_filter=horario_inicio_filter,
        tipo_acao_filter=tipo_acao_filter,
        equipamento_filter=equipamento_filter,
        solicitante_filter=solicitante_filter,
        codigo_falha_filter=codigo_falha_filter,
        area_filter=area_filter  
    )

@app.route('/api/hotspots', methods=['GET'])
@login_required
def get_hotspots():
    hotspots = Hotspot.query.all()  
    result = []
    for hotspot in hotspots:
        posicao = int(hotspot.posicao)  
        result.append({
            'id': hotspot.id,
            'top': hotspot.top,
            'left': hotspot.left,
            'info': {
                'posicao': hotspot.posicao
            }
        })
    return jsonify(result)
@app.route('/api/afiacoes', methods=['GET'])
@login_required
def get_afiacoes():
    afiacoes = Afiação.query.all()
    result = []
    for af in afiacoes:
        result.append({
            'id': af.id,
            'posicao': af.posicao,
            'ferramenta': af.ferramenta,
            'lado': af.lado,
            'altura': af.altura,
            'folga': af.folga,
            'spacer': af.spacer,
            'data_troca': af.data_troca.strftime('%Y-%m-%d %H:%M:%S'),
            'dias_produzidos': af.dias_produzidos,
            'ferramenteiro': af.ferramenteiro
        })
    return jsonify(result)

@app.route('/api/historico/all', methods=['GET'])
@login_required
def get_all_historico():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 27  
        
        query = HistoricoTroca.query.order_by(HistoricoTroca.data.desc())

        
        if data_inicio := request.args.get('dataInicio'):
            query = query.filter(HistoricoTroca.data >= data_inicio)
        if data_fim := request.args.get('dataFim'):
            query = query.filter(HistoricoTroca.data <= data_fim)
        if posicao := request.args.get('posicao'):
            query = query.filter(HistoricoTroca.posicao == posicao)
        if operador := request.args.get('operador'):
            query = query.filter(HistoricoTroca.operador.ilike(f'%{operador}%'))

        
        total_registros = query.count()

        
        all_producao = [item.producao_atual for item in query.all() if item.producao_atual]
        if all_producao:
            all_producao.sort()
            mid = len(all_producao) // 2
            media_vida_util = (
                all_producao[mid] if len(all_producao) % 2 
                else (all_producao[mid-1] + all_producao[mid]) / 2
            )
        else:
            media_vida_util = 0

        
        data_limite = datetime.now() - timedelta(days=30)
        operadores_ativos = db.session.query(
            db.distinct(HistoricoTroca.operador)
        ).filter(
            HistoricoTroca.data >= data_limite
        ).count()

        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        
        items = []
        for item in pagination.items:
            try:
                data_formatada = item.data.strftime('%Y-%m-%d %H:%M') if item.data else None
            except Exception as e:
                logging.error(f"Erro ao formatar data: {e}")
                data_formatada = None

            items.append({
                'id': item.id,
                'posicao': item.posicao,
                'codigo': item.codigo,
                'operador': item.operador,
                'data': data_formatada,
                'vida_util': item.vida_util,
                'producao_atual': item.producao_atual
            })

        return jsonify({
            'success': True,
            'items': items,
            'total': total_registros,
            'pages': pagination.pages,
            'current_page': page,
            'media_vida_util': media_vida_util,
            'total_operadores': operadores_ativos
        })

    except Exception as e:
        logging.error(f"Erro ao buscar histórico: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'items': []
        }), 500

@app.route('/api/troca-ferramenta', methods=['POST'])
@login_required
def registrar_troca():
    try:
        data = request.json
        
        
        cache_key = f"troca_{data['posicao']}_{data['ferramenta_nova_id']}"
        if hasattr(app, '_troca_cache') and cache_key in app._troca_cache:
            print(f"Duplicata detectada para: {cache_key}")
            return jsonify({
                'success': True,
                'message': 'Troca já registrada recentemente.',
                'duplicate': True
            })

        if not hasattr(app, '_troca_cache'):
            app._troca_cache = {}

        
        app._troca_cache[cache_key] = datetime.now()

        
        now = datetime.now()
        app._troca_cache = {k: v for k, v in app._troca_cache.items() 
                           if (now - v).total_seconds() < 5}

        
        ferramenta_nova = db.session.get(Ferramenta, data['ferramenta_nova_id'])
        if not ferramenta_nova:
            return jsonify({'error': 'Ferramenta não encontrada'}), 404

        
        ferramenta_atual = Ferramenta.query.filter(
            Ferramenta.posicao == int(data['posicao']),
            Ferramenta.status == 'em_uso',
            db.or_(
                db.func.substr(Ferramenta.codigo, 1, 3) == db.func.substr(ferramenta_nova.codigo, 1, 3),
                Ferramenta.tipo == ferramenta_nova.tipo
            )
        ).first()

        current_time = get_current_datetime()

        
        if ferramenta_atual:
            
            existing_record = HistoricoTroca.query.filter_by(
                posicao=int(data['posicao']),
                codigo=ferramenta_nova.codigo,
                data=current_time
            ).first()

            
            if not existing_record:
                nova_troca = HistoricoTroca(
                    posicao=int(data['posicao']),
                    codigo=ferramenta_nova.codigo,
                    operador=data['operador'],
                    vida_util=100000,
                    producao_atual=int(data['vida_util'])
                )
                db.session.add(nova_troca)
            
            ferramenta_atual.status = 'manutencao'
            ferramenta_atual.posicao = None

        
        ferramenta_nova.status = 'em_uso'
        ferramenta_nova.posicao = int(data['posicao'])
        ferramenta_nova.ultima_atualizacao = current_time
        
        db.session.commit()
        
        mensagem = 'Troca registrada com sucesso!' if ferramenta_atual else 'Ferramenta instalada com sucesso!'
        
        
        if cache_key in app._troca_cache:
            del app._troca_cache[cache_key]

        return jsonify({
            'success': True,
            'message': mensagem,
            'first_installation': ferramenta_atual is None
        })

    except Exception as e:
        db.session.rollback()
        
        if cache_key in app._troca_cache:
            del app._troca_cache
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas', methods=['GET'])
@login_required
def get_ferramentas():
    try:
        ferramentas = Ferramenta.query.all()
        return jsonify([{
            'id': f.id,
            'codigo': f.codigo,
            'tipo': f.tipo,
            'status': f.status,
            'descricao': f.descricao,
            'dimensao_metrica': f.dimensao_metrica,
            'dimensao_polegada': f.dimensao_polegada,
            'sufixo': f.sufixo,
            'posicao': f.posicao,
            'ultima_atualizacao': f.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
        } for f in ferramentas])
    except Exception as e:
        print(f"Erro ao buscar ferramentas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas/buscar/<codigo>', methods=['GET'])
@login_required
def buscar_ferramenta(codigo):
    try:
        ferramenta = Ferramenta.query.filter_by(codigo=codigo).first()
        if ferramenta:
            return jsonify({
                'id': ferramenta.id,
                'codigo': ferramenta.codigo,
                'tipo': ferramenta.tipo,
                'status': ferramenta.status,
                'ultima_atualizacao': ferramenta.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
            })
        return jsonify({'message': 'Ferramenta não encontrada'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas', methods=['POST'])
@login_required
def criar_ferramenta():
    try:
        data = request.json
        
        if not data.get('codigo'):
            return jsonify({'error': 'Código da ferramenta é obrigatório'}), 400

        ferramenta = Ferramenta.query.filter_by(codigo=data['codigo']).first()
        
        if ferramenta:
            
            ferramenta.tipo = data['tipo']
            if data.get('status'): 
                ferramenta.status = data['status']
            if data.get('posicao'): 
                ferramenta.posicao = data.get('posicao')
            ferramenta.ultima_atualizacao = datetime.utcnow()
            message = f'Ferramenta {data["codigo"]} atualizada com sucesso!'
        else:
            
            ferramenta = Ferramenta(
                codigo=data['codigo'],
                tipo=data['tipo'],
                status='disponivel',  
                posicao=None  
            )
            db.session.add(ferramenta)
            message = f'Ferramenta {data["codigo"]} cadastrada com sucesso!'

        db.session.commit()
        return jsonify({
            'success': True, 
            'message': message,
            'ferramenta': {
                'id': ferramenta.id,
                'codigo': ferramenta.codigo,
                'tipo': ferramenta.tipo,
                'status': ferramenta.status
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/get_filter_options', methods=['GET'])
@login_required
def get_filter_options():
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    
    try:
        # Add the area filter condition based on user's role
        where_clause = "1=1"
        params = []
        
        if current_user.area != 'supervisor':
            where_clause += " AND area = ?"
            params.append(current_user.area)

        queries = {
            'nomes': f"SELECT DISTINCT nome FROM historico WHERE {where_clause} ORDER BY nome",
            'solicitantes': f"SELECT DISTINCT solicitante FROM historico WHERE {where_clause} AND solicitante IS NOT NULL ORDER BY solicitante",
            'codigos_falha': f"SELECT DISTINCT codigo_falha FROM historico WHERE {where_clause} AND codigo_falha IS NOT NULL ORDER BY codigo_falha",
            'equipamentos': f"SELECT DISTINCT equipamento FROM historico WHERE {where_clause} AND equipamento IS NOT NULL ORDER BY equipamento",
            'tipos_acao': f"SELECT DISTINCT tipo_acao FROM historico WHERE {where_clause} AND tipo_acao IS NOT NULL ORDER BY tipo_acao"
        }
        
        result = {}
        for key, query in queries.items():
            c.execute(query, params)
            result[key] = [item[0] for item in c.fetchall() if item[0]]
            
        return jsonify(result)
        
    except Exception as e:
        logging.error(f"Error getting filter options: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/ferramentas/posicao/<int:posicao>', methods=['GET'])
@login_required
def get_ferramenta_posicao(posicao):
    try:
        
        tipo = request.args.get('tipo')
        if not tipo:
            return jsonify({'error': 'Tipo de ferramenta não especificado'}), 400

        
        tipo_codigo = next((code for code, name in AVAILABLE_TOOL_TYPES.items() 
                          if name.upper() == tipo.upper()), None)
        tipo_completo = AVAILABLE_TOOL_TYPES.get(tipo_codigo, tipo)

        
        ferramenta = Ferramenta.query.filter(
            Ferramenta.posicao == posicao,
            Ferramenta.status == 'em_uso',
            db.or_(
                db.func.substr(Ferramenta.codigo, 1, 3) == tipo_codigo if tipo_codigo else False,
                db.func.upper(Ferramenta.tipo) == tipo_completo.upper() if tipo_completo else False,
                db.func.upper(Ferramenta.tipo) == tipo.upper()
            )
        ).first()

        if ferramenta:
            ultima_troca = HistoricoTroca.query.filter_by(
                codigo=ferramenta.codigo,
                posicao=posicao
            ).order_by(HistoricoTroca.data.desc()).first()

            response_data = {
                'id': ferramenta.id,
                'codigo': ferramenta.codigo,
                'tipo': ferramenta.tipo,
                'status': ferramenta.status,
                'posicao': ferramenta.posicao,
                'operador': ultima_troca.operador if ultima_troca else current_user.nome,
                'data_ultima_troca': ultima_troca.data.strftime('%d/%m/%Y %H:%M') if ultima_troca else None,
                'ultima_atualizacao': ferramenta.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
            }
            return jsonify(response_data)
        
        return jsonify({
            'message': f'Nenhuma ferramenta do tipo {tipo_completo} instalada na posição {posicao}'
        }), 404

    except Exception as e:
        print(f"Erro ao buscar ferramenta na posição: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas/disponiveis', methods=['GET'])
@login_required
def get_ferramentas_disponiveis():
    try:
        ferramentas = Ferramenta.query.filter_by(status='disponivel').all()
        return jsonify([{
            'id': f.id,
            'codigo': f.codigo,
            'tipo': f.tipo,
            'ultima_atualizacao': f.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
        } for f in ferramentas])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas/disponiveis/<tipo>', methods=['GET'])
@login_required
def get_ferramentas_disponiveis_por_tipo(tipo):
    try:
        
        tipo_completo = AVAILABLE_TOOL_TYPES.get(tipo, tipo)
        tipo_codigo = next((code for code, name in AVAILABLE_TOOL_TYPES.items() if name == tipo), tipo)

        print(f"Buscando ferramentas do tipo: {tipo_completo} ou {tipo_codigo}")  

        
        ferramentas = Ferramenta.query.filter(
            Ferramenta.status == 'disponivel',
            db.or_(
                Ferramenta.tipo.ilike(f'%{tipo_completo}%'),
                Ferramenta.tipo.ilike(f'%{tipo}%'),
                Ferramenta.tipo.ilike(f'%{tipo_codigo}%')
            )
        ).all()

        print(f"Ferramentas encontradas: {len(ferramentas)}")  
        
        return jsonify([{
            'id': f.id,
            'codigo': f.codigo,
            'tipo': f.tipo,
            'ultima_atualizacao': f.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
        } for f in ferramentas])
    except Exception as e:
        print(f"Erro ao buscar ferramentas disponíveis: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ferramentas/<int:id>/posicao', methods=['PATCH'])
@login_required
def atualizar_posicao_ferramenta(id):
    try:
        data = request.json
        nova_posicao = int(data['posicao'])
        
        
        if nova_posicao < 1 or nova_posicao > 27:
            return jsonify({'error': f'Posição {nova_posicao} inválida. Deve estar entre 1 e 27.'}), 400

        
        ferramenta_nova = db.session.get(Ferramenta, id)
        if not ferramenta_nova:
            return jsonify({'error': 'Ferramenta não encontrada'}), 404

        
        ferramenta_atual = Ferramenta.query.filter_by(
            posicao=nova_posicao,
            status='em_uso'
        ).first()

        
        if ferramenta_atual:
            
            manutencao = ManutencaoFerramenta(
                ferramenta_id=ferramenta_atual.id,
                motivo_entrada=f"Substituição automática - Trocada por {ferramenta_nova.codigo}",
                operador_entrada=current_user.nome,
                status='pendente'
            )
            db.session.add(manutencao)
            
            
            ferramenta_atual.status = 'manutencao'
            ferramenta_atual.posicao = None
            
        
        ferramenta_nova.posicao = nova_posicao
        ferramenta_nova.status = 'em_uso'
        ferramenta_nova.ultima_atualizacao = datetime.utcnow()

        
        db.session.commit()

        return jsonify({
            'success': True, 
            'message': 'Posição atualizada com sucesso',
            'tool': {
                'id': ferramenta_nova.id,
                'codigo': ferramenta_nova.codigo,
                'tipo': ferramenta_nova.tipo,
                'posicao': ferramenta_nova.posicao,
                'status': ferramenta_nova.status
            },
            'old_tool': {
                'id': ferramenta_atual.id,
                'codigo': ferramenta_atual.codigo,
                'status': 'manutencao'
            } if ferramenta_atual else None
        })

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao atualizar posição: {str(e)}")
        return jsonify({'error': f'Erro ao atualizar posição: {str(e)}'}), 500


@app.route('/api/facas', methods=['GET'])
@login_required
def get_facas():
    try:
        posicao = request.args.get('posicao', type=int)
        query = Faca.query

        if posicao:
            
            query = query.filter_by(posicao=posicao)
        
        
        subquery = db.session.query(
            Faca.posicao,
            Faca.lado,
            db.func.max(Faca.data_troca).label('max_data')
        ).group_by(Faca.posicao, Faca.lado).subquery()

        facas_atuais = query.join(
            subquery,
            db.and_(
                Faca.posicao == subquery.c.posicao,
                Faca.lado == subquery.c.lado,
                Faca.data_troca == subquery.c.max_data
            )
        ).all()

        resultado = []
        for faca in facas_atuais:
            resultado.append({
                'id': faca.id,
                'posicao': faca.posicao,
                'ferramenta': faca.ferramenta,
                'lado': faca.lado,
                'altura': faca.altura,
                'folga': faca.folga,
                'spacer': faca.spacer,
                'data_troca': faca.data_troca.isoformat() if faca.data_troca else None,
                'dias_produzidos': faca.dias_produzidos,
                'utilizador': faca.utilizador
            })

        return jsonify(resultado)
    except Exception as e:
        print(f"Erro ao buscar facas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/facas', methods=['POST'])
@login_required
def add_faca():
    try:
        data = request.json
        posicao = int(data['posicao'])
        lado = data['lado']
        
        
        posicoes_apenas_lado_a = [5, 6, 13, 17]
        if posicao in posicoes_apenas_lado_a and lado != 'A':
            return jsonify({'error': 'Esta posição aceita apenas lado A'}), 400

        
        faca_atual = Faca.query.filter_by(
            posicao=posicao,
            lado=lado
        ).order_by(Faca.data_troca.desc()).first()

        
        nova_faca = Faca(
            posicao=posicao,
            ferramenta=data['ferramenta'],
            lado=lado,
            altura=float(data['altura']),
            folga=float(data['folga']),
            spacer=data.get('spacer'),
            utilizador=data['utilizador'],
            data_troca=datetime.utcnow(),
            dias_produzidos=0
        )
        
        
        if faca_atual:
            historico = HistoricoFacas(
                posicao=posicao,
                ferramenta_anterior=faca_atual.ferramenta,
                ferramenta_nova=data['ferramenta'],
                lado=lado,
                altura=float(data['altura']),
                folga=float(data['folga']),
                spacer=data.get('spacer'),
                utilizador=data['utilizador']
            )
            db.session.add(historico)

        
        db.session.add(nova_faca)
        db.session.commit()

        return jsonify({
            'message': 'Faca atualizada com sucesso!',
            'id': nova_faca.id
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao adicionar faca: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/facas/historico/<int:posicao>', methods=['GET'])
@login_required
def get_historico_facas(posicao):
    try:
        
        historico = HistoricoFacas.query.filter_by(posicao=posicao)\
            .order_by(HistoricoFacas.data_troca.desc())\
            .all()

        
        resultado = [{
            'id': h.id,
            'posicao': h.posicao,
            'ferramenta_anterior': h.ferramenta_anterior,
            'ferramenta_nova': h.ferramenta_nova,
            'lado': h.lado,
            'altura': h.altura,
            'folga': h.folga,
            'spacer': h.spacer,
            'data_troca': h.data_troca.isoformat(),
            'utilizador': h.utilizador
        } for h in historico]

        print(f"Histórico para posição {posicao}:", resultado)  
        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar histórico: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/adicionar_matricula', methods=['GET', 'POST'])
@login_required
def adicionar_matricula():
    print(f"Usuário atual: {current_user.get_id()}")  
    
    user_id = int(current_user.get_id())  
    
    if user_id not in [0,1,2]:  
        flash('Acesso negado. Apenas administradores podem adicionar matrículas.', 'danger')
        return redirect(url_for('relatorio'))

    if request.method == 'POST':
        nome = request.form['nome']  
        matricula = request.form['matricula']
        area = request.form['area']  

        conn = sqlite3.connect('relatorio_diario.db')
        c = conn.cursor()
        
        
        c.execute("SELECT matricula FROM admin WHERE matricula = ?", (matricula,))
        usuario_existente = c.fetchone()

        if usuario_existente:
            flash('Matricula inválida.', 'danger')
        else:
            
            senha_atual = f'canpack.{datetime.now().year}'
            senha_hash = generate_password_hash(senha_atual)
            c.execute(
                "INSERT INTO admin (nome, matricula, senha, area) VALUES (?, ?, ?, ?)",
                (nome, matricula, senha_hash, area)
            )
            conn.commit()
            flash(f'Usuário adicionado com sucesso à área de {area}!', 'success')

        conn.close()
    
    return render_template('adicionar_matricula.html')


@app.route('/relatorio', methods=['GET', 'POST'])
@login_required
def relatorio():
    if request.method == 'POST':
        
        nome = request.form.get('nome', '').strip()
        tipo_acao = request.form.get('tipo_acao', '').strip()
        equipamento = request.form.get('equipamento', '').strip()
        solicitante = request.form.get('solicitante', '').strip()
        codigo_falha = request.form.get('codigo_falha', '').strip()
        causa_encontrada = request.form.get('causa_encontrada', '').strip()
        trabalho_executado_original = request.form.get('trabalho_executado', '').strip()
        comentario = request.form.get('comentario', '').strip()
        data = request.form.get('data', '').strip()
        horario_inicio = request.form.get('horario_inicio', '').strip()
        horario_termino = request.form.get('horario_termino', '').strip()
        foto = request.files.get('foto')
        data_alterada_manual = request.form.get('data_alterada_manual', 'false') == 'true'

        
        erros = []
        if not nome:
            erros.append("O campo Nome é obrigatório.")
        if not data:
            erros.append("O campo Data é obrigatório.")
        if not horario_inicio:
            erros.append("O campo Horário de Início é obrigatório.")
        if not horario_termino:
            erros.append("O campo Horário de Término é obrigatório.")
        if equipamento in ['Interno', 'Outros'] and not comentario:
            erros.append("O campo Comentário é obrigatório para os equipamentos 'Interno' ou 'Outros'.")
        if data:
            try:
                data_selecionada = datetime.strptime(data, '%Y-%m-%d').date()
                data_hoje = datetime.now().date()
                diferenca_dias = (data_hoje - data_selecionada).days
                if data_selecionada > data_hoje:
                    erros.append("Não é permitido registrar atividades em datas futuras.")
                elif diferenca_dias > 3:
                    data_limite = (data_hoje - timedelta(days=3)).strftime('%d/%m/%Y')
                    erros.append(f"Não é permitido registrar atividades com mais de 3 dias de retroatividade. A data mais antiga permitida é {data_limite}.")
            except ValueError:
                erros.append("Formato de data inválido.")
        if foto and not allowed_file(foto.filename):
            erros.append("Extensão do arquivo não permitida.")
        if erros:
            for erro in erros:
                flash(erro, 'danger')
            return render_template('relatorio.html')
        foto_path = None
        if foto:
            foto_nome = secure_filename(foto.filename)
            foto_path = os.path.join(app.config['UPLOAD_FOLDER'], foto_nome)
            foto.save(foto_path)
        try:
            tempo_inicio = int(horario_inicio.split(":")[0]) * 60 + int(horario_inicio.split(":")[1])
            tempo_termino = int(horario_termino.split(":")[0]) * 60 + int(horario_termino.split(":")[1])
            eficiencia = max((tempo_termino - tempo_inicio) / 60, 0) * 100
        except Exception:
            eficiencia = 0
        area = request.form.get('area')
        if current_user.area != 'supervisor':
            area = current_user.area
        elif area not in ['tampas', 'latas']:
            area = 'latas'  
        conn = sqlite3.connect('relatorio_diario.db')
        c = conn.cursor()
        
        try:
            data_atual = datetime.now().strftime('%Y-%m-%d')
            atividades = [
                atividade.strip() for atividade in 
                re.split(r'\s*[;+]\s*', trabalho_executado_original)
                if atividade.strip()
            ]
            if not atividades:
                atividades = [trabalho_executado_original]
            is_split = len(atividades) > 1
            for trabalho_executado in atividades:
                c.execute('''INSERT INTO historico (
                    data, nome, tipo_acao, equipamento, solicitante, codigo_falha,
                    causa_encontrada, trabalho_executado, comentario, 
                    horario_inicio, horario_termino, foto, eficiencia, area
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                    (data, nome, tipo_acao, equipamento, solicitante, codigo_falha,
                     causa_encontrada, trabalho_executado, comentario,
                     horario_inicio, horario_termino, foto_path, eficiencia, area))
                
                registro_id = c.lastrowid
                log_messages = []
                backup_data = data
                if data_alterada_manual and data != data_atual:
                    log_messages.append(f"Data alterada manualmente para {data}")
                    backup_data = data_atual
                if is_split:
                    log_messages.append(f"Atividade desmembrada de: '{trabalho_executado_original}'")
                if log_messages:
                    full_log_message = f"{current_user.nome} ({'; '.join(log_messages)})"
                    c.execute('''INSERT INTO historico_backup (
                        id_original, data, nome, tipo_acao, equipamento, solicitante,
                        codigo_falha, causa_encontrada, trabalho_executado, comentario,
                        horario_inicio, horario_termino, foto, eficiencia,
                        editado_por, data_edicao
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))''', 
                        (registro_id, backup_data, nome, tipo_acao, equipamento, solicitante, codigo_falha,
                         causa_encontrada, trabalho_executado, comentario,
                         horario_inicio, horario_termino, foto_path, eficiencia,
                         full_log_message))
            conn.commit()
            flash('Relatório adicionado com sucesso!', 'success')
            if is_split:
                flash('A atividade foi desmembrada em várias entradas.', 'info')
            elif data_alterada_manual:
                 flash(f'Data alterada manualmente de {data_atual} para {data}.', 'info')
        except sqlite3.IntegrityError as e:
            conn.rollback()
            flash('Erro ao salvar relatório: valor inválido para área.', 'danger')
            logging.error(f"Erro de integridade: {str(e)}")
        except Exception as e:
            conn.rollback()
            flash('Erro ao salvar relatório.', 'danger')
            logging.error(f"Erro ao salvar relatório: {str(e)}")
        finally:
            conn.close()
        return redirect(url_for('relatorio'))
    return render_template('relatorio.html')


@app.route('/download_excel', methods=['GET', 'POST'])
@login_required
def download_excel():
    search_filter = request.args.get('search')
    data_filter = request.args.get('data')
    nome_filter = request.args.get('nome')
    horario_inicio_filter = request.args.get('horario_inicio')
    tipo_acao_filter = request.args.get('tipo_acao')
    equipamento_filter = request.args.get('equipamento')
    solicitante_filter = request.args.get('solicitante')
    codigo_falha_filter = request.args.get('codigo_falha')
    all_data = request.args.get('allData')  
    logging.debug(f"Filtros recebidos: search={search_filter}, data={data_filter}, "
                  f"nome={nome_filter}, horario_inicio={horario_inicio_filter}, "
                  f"tipo_acao={tipo_acao_filter}, equipamento={equipamento_filter}, "
                  f"solicitante={solicitante_filter}, codigo_falha={codigo_falha_filter}, allData={all_data}")
    conn = sqlite3.connect('relatorio_diario.db')
    query = """
        SELECT
            id,
            data,
            nome,
            horario_inicio,
            tipo_acao,
            equipamento,
            solicitante,
            codigo_falha,
            causa_encontrada,
            trabalho_executado,
            comentario,
            foto,
            area
        FROM historico
        WHERE 1=1
    """
    params = []
    area_filter = request.args.get('area')
    if current_user.area == 'supervisor':
        if area_filter:  
            query += " AND area = ?"
            params.append(area_filter)
    else:  
        query += " AND area = ?"
        params.append(current_user.area)
    if not all_data:
        if search_filter:
            query += " AND (id = ? OR nome LIKE ?)"
            params.append(search_filter)
            params.append(f'%{search_filter}%')
        if data_filter:
            try:
                data_filter = datetime.strptime(data_filter, '%Y-%m-%d').strftime('%Y-%m-%d')
                query += " AND DATE(data) = ?"
                params.append(data_filter)
            except ValueError:
                pass  
        if nome_filter:
            query += " AND nome LIKE ?"
            params.append(f'%{nome_filter}%')
        if horario_inicio_filter:
            query += " AND horario_inicio = ?"
            params.append(horario_inicio_filter)
        if tipo_acao_filter:
            query += " AND tipo_acao LIKE ?"
            params.append(f'%{tipo_acao_filter}%')
        if equipamento_filter:
            query += " AND equipamento LIKE ?"
            params.append(f'%{equipamento_filter}%')
        if solicitante_filter:
            query += " AND solicitante LIKE ?"
            params.append(f'%{solicitante_filter}%')
        if codigo_falha_filter:
            query += " AND codigo_falha LIKE ?"
            params.append(f'%{codigo_falha_filter}%')

    
    logging.debug(f"Download Query: {query}")
    logging.debug(f"Download Params: {params}")
    logging.debug(f"User area: {current_user.area}")

    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    
    if df.empty:
        flash('Nenhum dado disponível para gerar o relatório.', 'danger')
        return redirect(url_for('historico_acoes'))

    
    for idx, row in df.iterrows():
        foto_name = row['foto']
        if foto_name:
            foto_path_completo = os.path.abspath(os.path.join(foto_name))
            if os.path.exists(foto_path_completo):
                df.at[idx, 'foto'] = f'=HYPERLINK("{foto_path_completo}", "Clique para ver imagem")'
            else:
                df.at[idx, 'foto'] = f'=HYPERLINK("{foto_path_completo}", "Clique para ver imagem (Imagem não encontrada)")'
        else:
            df.at[idx, 'foto'] = 'Sem foto'

    
    df['data'] = pd.to_datetime(df['data'], errors='coerce')
    df['ano'] = df['data'].dt.year
    df['mes'] = df['data'].dt.strftime('%B')

    
    base_dir = 'relatorio'
    if (os.path.exists(base_dir)):
        os.makedirs(base_dir)

    excel_path = os.path.join(base_dir, 'relatorio_filtrado.xlsx')
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.drop(columns=['ano', 'mes'], inplace=True)
        df.to_excel(writer, sheet_name='Relatório', index=False)

    
    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True, download_name='relatorio_filtrado.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        flash('Erro ao gerar o relatório. Tente novamente.', 'danger')
        return redirect(url_for('historico_acoes'))


@app.route('/historico_todas_edicoes', methods=['GET'])
@login_required
def historico_todas_edicoes():
    if current_user.area != 'supervisor':
        flash('Acesso negado. Apenas supervisores podem ver o histórico completo de edições.', 'danger')
        return redirect(url_for('historico_acoes'))

    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    
    page = request.args.get('page', 1, type=int)
    per_page = 20  
    
    
    c.execute("SELECT COUNT(*) FROM historico_backup")
    total_records = c.fetchone()[0]
    total_pages = ceil(total_records / per_page)
    
    query = """
        SELECT 
            hb.id,
            hb.id_original,
            hb.data_edicao,
            hb.editado_por,
            hb.data,
            hb.nome,
            hb.tipo_acao,
            hb.equipamento,
            hb.area,
            h.area as area_atual
        FROM historico_backup hb
        LEFT JOIN historico h ON h.id = hb.id_original
        ORDER BY hb.data_edicao DESC
        LIMIT ? OFFSET ?
    """
    
    offset = (page - 1) * per_page
    c.execute(query, (per_page, offset))
    registros = c.fetchall()
    conn.close()

    return render_template(
        'historico_edicoes.html',
        registros=registros,
        page=page,
        total_pages=total_pages,
        total_records=total_records
    )

@app.route('/historico_edicoes/<int:registro_id>', methods=['GET'])
@login_required
def get_edit_details(registro_id):
    if current_user.area != 'supervisor':
        return jsonify({'error': 'Acesso não autorizado'}), 403

    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    
    
    c.execute("""
        SELECT 
            id,
            data,
            nome,
            horario_inicio,
            tipo_acao,
            equipamento,
            solicitante,
            codigo_falha,
            causa_encontrada,
            trabalho_executado,
            comentario,
            foto,
            area
        FROM historico 
        WHERE id = ?
    """, (registro_id,))
    current_version = c.fetchone()
    
    
    c.execute("""
        SELECT 
            id,
            data,
            nome,
            horario_inicio,
            tipo_acao,
            equipamento,
            solicitante,
            codigo_falha,
            causa_encontrada,
            trabalho_executado,
            comentario,
            foto,
            area,
            editado_por,
            data_edicao
        FROM historico_backup 
        WHERE id_original = ? 
        ORDER BY data_edicao DESC
    """, (registro_id,))
    backups = c.fetchall()
    
    
    current_columns = [
        'id', 'data', 'nome', 'horario_inicio', 'tipo_acao', 'equipamento', 
        'solicitante', 'codigo_falha', 'causa_encontrada', 'trabalho_executado', 
        'comentario', 'foto', 'area'
    ]
    
    backup_columns = [
        'id', 'data', 'nome', 'horario_inicio', 'tipo_acao', 'equipamento', 
        'solicitante', 'codigo_falha', 'causa_encontrada', 'trabalho_executado', 
        'comentario', 'foto', 'area', 'editado_por', 'data_edicao'
    ]
    
    
    response = {
        'current': dict(zip(current_columns, current_version)) if current_version else None,
        'history': [dict(zip(backup_columns, backup)) for backup in backups]
    }
    
    conn.close()
    return jsonify(response)

@app.route('/api/manutencao', methods=['POST'])
@login_required
def registrar_manutencao():
    try:
        data = request.json
        ferramenta_id = data.get('ferramenta_id')
        motivo = data.get('motivo')

        if not ferramenta_id  or not motivo:
            return jsonify({'error': 'Dados incompletos: ferramenta_id e motivo são obrigatórios'}), 400

        ferramenta = db.session.get(Ferramenta, ferramenta_id)
        if not ferramenta:
            return jsonify({'error': 'Ferramenta não encontrada'}), 404

        manutencao = ManutencaoFerramenta(
            ferramenta_id=ferramenta_id,
            motivo_entrada=motivo,
            operador_entrada=current_user.nome
        )

        ferramenta.status = 'manutencao'

        db.session.add(manutencao)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Ferramenta enviada para manutenção'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/api/manutencao/<int:ferramenta_id>/concluir', methods=['POST'])
@login_required
def concluir_manutencao(ferramenta_id):
    try:
        data = request.json
        ferramenta = db.session.get(Ferramenta, ferramenta_id)
        
        if not ferramenta:
            return jsonify({'error': 'Ferramenta não encontrada'}), 404

        if data['status'] == 'descartada':
            
            descarte = DescarteFerramenta(
                ferramenta_id=ferramenta.id,
                codigo=ferramenta.codigo,
                operador=data.get('responsavel', 'Não informado'),
                motivo=data.get('descricao', 'Sem motivo informado'),
                data_descarte=get_current_datetime()
            )
            db.session.add(descarte)
            ferramenta.status = 'descartada'
        else:
            ferramenta.status = 'disponivel'

        
        manutencao = ManutencaoFerramenta.query.filter_by(
            ferramenta_id=ferramenta_id,
            status='pendente'
        ).order_by(ManutencaoFerramenta.data_entrada.desc()).first()

        if manutencao:
            manutencao.status = 'concluido' if data['status'] == 'disponivel' else 'descartado'
            manutencao.data_saida = get_current_datetime()
            manutencao.operador_saida = data.get('responsavel')
            manutencao.descricao_manutencao = data.get('descricao')
            if data['status'] == 'descartada':
                manutencao.motivo_descarte = data.get('descricao')

        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Manutenção concluída com sucesso',
            'status': data['status']
        })

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao concluir manutenção: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ferramentas/manutencao/<int:ferramenta_id>', methods=['GET'])
@login_required
def get_ferramenta_manutencao(ferramenta_id):
    try:
        ferramenta = db.session.get(Ferramenta, ferramenta_id)
        if not ferramenta:
            return jsonify({'error': 'Ferramenta não encontrada'}), 404

        
        manutencoes_pendentes = ManutencaoFerramenta.query.filter_by(
            ferramenta_id=ferramenta_id,
            status='pendente'
        ).all()

        if manutencoes_pendentes:
            for manutencao in manutencoes_pendentes:
                manutencao.status = 'em uso'  
                ferramenta.status = 'disponível'  

        db.session.commit()

        return jsonify({
            'codigo': ferramenta.codigo,
            'tipo': ferramenta.tipo,
            'status': ferramenta.status,
            'manutencoes_pendentes': len(manutencoes_pendentes)  
            
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500
    
@app.route('/api/ferramentas/descartadas', methods=['GET'])
def get_ferramentas_descartadas():
    try:
        
        descartes = DescarteFerramenta.query.order_by(
            DescarteFerramenta.data_descarte.desc()
        ).all()

        resultado = []
        for d in descartes:
            descarte_data = {
                'id': d.id,
                'codigo': d.codigo,
                'motivo': d.motivo,
                'operador': d.operador,
                'data_descarte': None
            }
            
            
            if d.data_descarte:
                try:
                    
                    data_local = d.data_descarte.astimezone()
                    descarte_data['data_descarte'] = data_local.strftime('%d/%m/%Y %H:%M')
                except Exception as date_error:
                    print(f"Erro ao formatar data: {date_error}")
                    descarte_data['data_descarte'] = str(d.data_descarte)
            
            resultado.append(descarte_data)

        return jsonify(resultado)
    except Exception as e:
        print(f"Erro ao buscar descartes: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/historico_descarte')
@login_required
def historico_descarte():
    return render_template('historico_descarte.html')

@app.route('/historico_trocas')
@login_required
def historico_trocas():
    return render_template('historico_trocas.html')

@app.errorhandler(401)
def unauthorized_error(error):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Unauthorized access'}), 401
    return redirect(url_for('login'))




EXCEL_STYLES = {
    'header': {
        'fill': PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
        'font': Font(color="FFFFFF", bold=True, size=12),
        'border': Border(
            left=Side(style='medium', color='FFFFFF'),
            right=Side(style='medium', color='FFFFFF'),
            top=Side(style='medium', color='FFFFFF'),
            bottom=Side(style='medium', color='FFFFFF')
        ),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'cell': {
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'status_colors': {
        'success': PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        'warning': PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
        'danger': PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    }
}

def gerar_excel_historico(dados, filtros=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Trocas"

    
    ws.column_dimensions['A'].width = 10  
    ws.column_dimensions['B'].width = 15  
    ws.column_dimensions['C'].width = 25  
    ws.column_dimensions['D'].width = 20  
    ws.column_dimensions['E'].width = 15  
    ws.column_dimensions['F'].width = 15  
    ws.column_dimensions['G'].width = 15  

    
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório de Histórico de Trocas'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = EXCEL_STYLES['header']['alignment']

    
    row = 2

    
    headers = ['Posição', 'Código', 'Operador', 'Data', 'Vida Útil', 'Produção', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = EXCEL_STYLES['header']['fill']
        cell.font = EXCEL_STYLES['header']['font']
        cell.border = EXCEL_STYLES['header']['border']
        cell.alignment = EXCEL_STYLES['header']['alignment']

    
    for item in dados:
        row += 1
        ws.cell(row=row, column=1, value=item['posicao'])
        ws.cell(row=row, column=2, value=item['codigo'])
        ws.cell(row=row, column=3, value=item['operador'])
        ws.cell(row=row, column=4, value=item['data'])
        ws.cell(row=row, column=5, value=item['vida_util'])
        ws.cell(row=row, column=6, value=item['producao_atual'])
        
        
        percentual = (item['producao_atual'] / item['vida_util']) * 100
        status_cell = ws.cell(row=row, column=7)
        
        if percentual < 60:
            status_cell.fill = EXCEL_STYLES['status_colors']['success']
            status_cell.value = "Normal"
        elif percentual < 85:
            status_cell.fill = EXCEL_STYLES['status_colors']['warning']
            status_cell.value = "Atenção"
        else:
            status_cell.fill = EXCEL_STYLES['status_colors']['danger']
            status_cell.value = "Crítico"

        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = EXCEL_STYLES['cell']['border']
            cell.alignment = EXCEL_STYLES['cell']['alignment']

    return wb

def enviar_email(subject, body, recipient_email, attachment=None, attachment_filename=None, embedded_images=None):
    from email.mime.image import MIMEImage

    msg = MIMEMultipart('related')
    msg['From'] = config.SMTP_USERNAME
    if isinstance(recipient_email, list):
        msg['To'] = ", ".join(recipient_email)
    else:
        msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    if attachment and attachment_filename:
        from email.mime.base import MIMEBase
        from email import encoders
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{attachment_filename}"')
        msg.attach(part)

    if embedded_images:
        for img_data in embedded_images:
            try:
                with open(img_data['path'], 'rb') as f:
                    img = MIMEImage(f.read())
                    img.add_header('Content-ID', f"<{img_data['cid']}>")
                    msg.attach(img)
            except FileNotFoundError:
                logging.warning(f"Image file not found: {img_data['path']}")

    try:
        server = smtplib.SMTP(config.SMTP_SERVER, config.SMTP_PORT)
        server.starttls()
        server.login(config.SMTP_USERNAME, config.SMTP_PASSWORD)
        server.sendmail(config.SMTP_USERNAME, recipient_email, msg.as_bytes())
        server.quit()
        if isinstance(recipient_email, list):
            logging.info(f"Email sent successfully to {', '.join(recipient_email)}")
        else:
            logging.info(f"Email sent successfully to {recipient_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def enviar_email_relatorio_diario(turno=None):
    
    conn = sqlite3.connect('relatorio_diario.db')
    c = conn.cursor()
    now = datetime.now()
    data_hoje = now.strftime('%Y-%m-%d')
    data_ontem = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    # Definir faixas de horário para cada turno
    if turno == 'noite':
        # Turno noite: 19:00 do dia anterior até 06:59 do dia atual
        where = "( (data = ? AND horario_inicio >= '19:00') OR (data = ? AND horario_inicio <= '06:59') )"
        params = (data_ontem, data_hoje)
    elif turno == 'manha':
        # Turno manhã: 07:00 até 18:59 do dia atual
        where = "(data = ? AND horario_inicio >= '07:00' AND horario_inicio <= '18:59')"
        params = (data_hoje,)
    else:
        where = "(data = ?)"
        params = (data_hoje,)
    if turno == 'noite':
        areas_para_enviar = ['latas']
    elif turno == 'manha':
        areas_para_enviar = ['tampas']
    else:
        areas_para_enviar = ['tampas', 'latas']

    # evitar múltiplos envios da mesma área nesta execução
    sent_areas = set()

    for area in areas_para_enviar:
        if area in sent_areas:
            logging.info(f"[EMAIL] Área '{area}' já enviada nesta execução. Pulando.")
            continue

        try:
            c.execute(f"SELECT id, horario_inicio, tipo_acao, equipamento, nome, trabalho_executado, comentario, foto FROM historico WHERE area = ? AND {where} ORDER BY horario_inicio", (area, *params))
            rows = c.fetchall()
            
            data_for_template = []
            embedded_images = []
            
            for row in rows:
                row_list = list(row)
                if row_list[7] and os.path.exists(row_list[7]):
                    image_path = row_list[7]
                    cid = f"image_{uuid.uuid4().hex}"
                    row_list[7] = cid
                    embedded_images.append({'path': image_path, 'cid': cid})
                else:
                    row_list[7] = None
                data_for_template.append(row_list)

            summary = defaultdict(int)
            for row in data_for_template:
                summary[row[2]] += 1

        except Exception as e:
            logging.error(f"Error processing {area} area: {e}")
            data_for_template = []
            embedded_images = []
            summary = defaultdict(int)

        area_name = 'Tampas' if area == 'tampas' else 'Latas'
        body = format_email_body(data_for_template, area_name, summary)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ações {area_name}"
        headers = ["ID", "Hora", "Tipo", "Máquina", "Responsável", "Atividade", "Detalhes", "Foto"]
        ws.append(headers)
        
        for row in data_for_template:
            ws.append(row)
            
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Get the recipient list for the current area and deduplicate
        recipients_area = config.RECIPIENT_EMAIL.get(area) or []
        recipients_supervisor = config.RECIPIENT_EMAIL.get('supervisor') or []
        # manter ordem e remover duplicatas
        combined = recipients_area + recipients_supervisor
        recipient_list = list(dict.fromkeys(combined))

        if not recipient_list:
            logging.warning(f"[EMAIL] Nenhum destinatário configurado para área '{area}'. Pulando envio.")
            sent_areas.add(area)
            continue

        logging.info(f"[EMAIL] Enviando relatório para área '{area}' ({area_name}) para {len(recipient_list)} destinatário(s): {recipient_list}")

        # Send email (apenas uma chamada por área)
        enviar_email(
            f"Relatório de Ações - {area_name} - {now.strftime('%d/%m/%Y %H:%M')}",
            body,
            recipient_list,
            attachment=excel_buffer,
            attachment_filename=f"relatorio_{area}_{now.strftime('%Y%m%d_%H%M')}.xlsx",
            embedded_images=embedded_images
        )

        # marcar como enviado para evitar duplicatas nesta execução
        sent_areas.add(area)

    conn.close()

def format_email_body(data, area_name, summary_data):
    with open('templates/SMTP/TEMPLATE_ENVIO.HTML', 'r', encoding='utf-8') as f:
        template_str = f.read()
    
    template = Template(template_str)
    
    current_datetime = get_current_datetime()
    generation_date = format_datetime(current_datetime, '%d/%m/%Y')
    generation_time = format_datetime(current_datetime, '%H:%M')

    return template.render(
        area_name=area_name, 
        data=data, 
        generation_date=generation_date,
        generation_time=generation_time,
        summary=summary_data,
        total_activities=len(data)
    )

@app.route('/api/historico/export', methods=['GET'])
@login_required
def exportar_historico():
    try:
        logging.info("Iniciando exportação de histórico")
        filtros = {
            'Data Início': request.args.get('dataInicio'),
            'Data Fim': request.args.get('dataFim'),
            'Posição': request.args.get('posicao'),
            'Operador': request.args.get('operador')
        }

        logging.info(f"Filtros recebidos: {filtros}")  

        
        query = HistoricoTroca.query.order_by(HistoricoTroca.data.desc())

        
        if filtros['Data Início']:
            query = query.filter(HistoricoTroca.data >= filtros['Data Início'])
        if filtros['Data Fim']:
            query = query.filter(HistoricoTroca.data <= filtros['Data Fim'])
        if filtros['Posição']:
            query = query.filter(HistoricoTroca.posicao == filtros['Posição'])
        if filtros['Operador']:
            query = query.filter(HistoricoTroca.operador.ilike(f"%{filtros['Operador']}%"))

        
        resultados = query.all()
        logging.info(f"Total de registros encontrados: {len(resultados)}")  

        if not resultados:
            return jsonify({'error': 'Nenhum dado encontrado para exportar'}), 404

        
        dados = [{
            'posicao': item.posicao,
            'codigo': item.codigo,
            'operador': item.operador,
            'data': item.data.strftime('%d/%m/%Y %H:%M') if item.data else None,
            'vida_util': item.vida_util,
            'producao_atual': item.producao_atual
        } for item in resultados]

        
        wb = gerar_excel_historico(dados, filtros)

        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        
        filename = f'historico_trocas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        logging.info(f"Arquivo gerado com sucesso: {filename}")  

        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error(f"Erro ao exportar histórico: {e}")
        return jsonify({
            'error': str(e),
            'detail': traceback.format_exc()
        }), 500

def gerar_excel_historico_acoes(dados, filtros=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Ações"

    
    colunas = {
        'A': ('ID', 8),
        'B': ('Data', 12),
        'C': ('Nome', 30),
        'D': ('Hora Início', 12),
        'E': ('Hora Término', 12),
        'F': ('Duração', 12),
        'G': ('Tipo de Ação', 25),
        'H': ('Equipamento', 25),
        'I': ('Solicitante', 20),
        'J': ('Código Falha', 15),
        'K': ('Causa encontrada', 40),
        'L': ('Trabalho executado', 40),
        'M': ('Comentário', 30),
        'N': ('Área', 12)
    }

    
    for col, (_, width) in colunas.items():
        ws.column_dimensions[col].width = width

    
    ultima_coluna = get_column_letter(len(colunas))
    ws.merge_cells(f'A1:{ultima_coluna}1')
    ws['A1'] = 'Relatório de Histórico de Ações'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = EXCEL_STYLES['header']['alignment']

    
    current_row = 2
    for col, (header, _) in enumerate(colunas.values(), 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = EXCEL_STYLES['header']['fill']
        cell.font = EXCEL_STYLES['header']['font']
        cell.border = EXCEL_STYLES['header']['border']
        cell.alignment = EXCEL_STYLES['header']['alignment']

    
    for row_num, item in enumerate(dados, start=current_row + 1):
        for col_num, value in enumerate(item, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = EXCEL_STYLES['cell']['border']
            cell.alignment = EXCEL_STYLES['cell']['alignment']
            
            
            if col_num in [11, 12]:  # Causa e Trabalho
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    return wb

@app.route('/export_excel_acoes', methods=['GET'])
@login_required
def export_excel_acoes():
    try:
        search_filter = request.args.get('search')
        data_filter = request.args.get('data')
        nome_filter = request.args.get('nome')
        horario_inicio_filter = request.args.get('horario_inicio')
        tipo_acao_filter = request.args.get('tipo_acao')
        equipamento_filter = request.args.get('equipamento')
        solicitante_filter = request.args.get('solicitante')
        codigo_falha_filter = request.args.get('codigo_falha')
        area_filter = request.args.get('area')

        
        conn = sqlite3.connect('relatorio_diario.db')
        query = """
            SELECT
                h.id,
                h.data,
                h.nome,
                h.horario_inicio,
                h.horario_termino,
                h.tipo_acao,
                h.equipamento,
                h.solicitante,
                h.codigo_falha,
                h.causa_encontrada,
                h.trabalho_executado,
                h.comentario,
                h.area
            FROM historico h
            WHERE 1=1
        """
        params = []
        if current_user.area == 'supervisor':
            if area_filter:
                query += " AND h.area = ?"
                params.append(area_filter)
        else:
            query += " AND h.area = ?"
            params.append(current_user.area)
        if search_filter:
            query += " AND (h.id = ? OR h.nome LIKE ?)"
            params.extend([search_filter, f'%{search_filter}%'])
        if data_filter:
            query += " AND DATE(h.data) = ?"
            params.append(data_filter)
        if nome_filter:
            query += " AND h.nome LIKE ?"
            params.append(f'%{nome_filter}%')
        if horario_inicio_filter:
            query += " AND h.horario_inicio = ?"
            params.append(horario_inicio_filter)
        if tipo_acao_filter:
            query += " AND h.tipo_acao LIKE ?"
            params.append(f'%{tipo_acao_filter}%')
        if equipamento_filter:
            query += " AND h.equipamento LIKE ?"
            params.append(f'%{equipamento_filter}%')
        if solicitante_filter:
            query += " AND h.solicitante LIKE ?"
            params.append(f'%{solicitante_filter}%')
        if codigo_falha_filter:
            query += " AND h.codigo_falha LIKE ?"
            params.append(f'%{codigo_falha_filter}%')
        query += " ORDER BY h.data DESC, h.horario_inicio DESC"
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        if df.empty:
            return jsonify({'error': 'Nenhum dado encontrado para exportar'}), 404

        def calculate_duration(row):
            if pd.isna(row['horario_inicio']) or pd.isna(row['horario_termino']):
                return None
            try:
                start = datetime.strptime(row['horario_inicio'], '%H:%M')
                end = datetime.strptime(row['horario_termino'], '%H:%M')
                if end < start:
                    end += timedelta(days=1)
                duration = end - start
                return str(duration)
            except (ValueError, TypeError):
                return None

        df['duracao'] = df.apply(calculate_duration, axis=1)
        df['data'] = pd.to_datetime(df['data']).dt.strftime('%d/%m/%Y')
        
        df = df[['id', 'data', 'nome', 'horario_inicio', 'horario_termino', 'duracao', 'tipo_acao', 'equipamento', 'solicitante', 'codigo_falha', 'causa_encontrada', 'trabalho_executado', 'comentario', 'area']]

        wb = gerar_excel_historico_acoes(df.values.tolist())
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'historico_acoes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )

    except Exception as e:
        logging.error(f"Erro ao exportar excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

def gerar_excel_descartes(dados, filtros=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Descartes"

    
    colunas = {
        'A': ('Código', 15),
        'B': ('Data Descarte', 20),
        'C': ('Operador', 25),
        'D': ('Motivo', 50),
    }

    
    for col, (_, width) in colunas.items():
        ws.column_dimensions[col].width = width
    ultima_coluna = get_column_letter(len(colunas))
    ws.merge_cells(f'A1:{ultima_coluna}1')
    ws['A1'] = 'Relatório de Histórico de Descartes'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = EXCEL_STYLES['header']['alignment']
    current_row = 2
    for col, (header, _) in enumerate(colunas.values(), 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = EXCEL_STYLES['header']['fill']
        cell.font = EXCEL_STYLES['header']['font']
        cell.border = EXCEL_STYLES['header']['border']
        cell.alignment = EXCEL_STYLES['header']['alignment']
    for row_num, item in enumerate(dados, start=current_row + 1):
        ws.cell(row=row_num, column=1, value=item['codigo'])
        ws.cell(row=row_num, column=2, value=item['data_descarte'])
        ws.cell(row=row_num, column=3, value=item['operador'])
        ws.cell(row=row_num, column=4, value=item['motivo'])

        
        for col in range(1, len(colunas) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = EXCEL_STYLES['cell']['border']
            cell.alignment = EXCEL_STYLES['cell']['alignment']
            if col == 4:  
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    return wb

@app.route('/api/descartes/export', methods=['GET'])
@login_required
def exportar_descartes():
    try:
        
        search_term = request.args.get('search', '').lower()
        data_inicial = request.args.get('dataInicial')
        data_final = request.args.get('dataFinal')

        
        query = DescarteFerramenta.query.order_by(DescarteFerramenta.data_descarte.desc())

        
        if search_term:
            query = query.filter(db.or_(
                DescarteFerramenta.codigo.ilike(f'%{search_term}%'),
                DescarteFerramenta.operador.ilike(f'%{search_term}%'),
                DescarteFerramenta.motivo.ilike(f'%{search_term}%')
            ))
        
        if data_inicial:
            query = query.filter(DescarteFerramenta.data_descarte >= data_inicial)
        if data_final:
            query = query.filter(DescarteFerramenta.data_descarte <= data_final)
        descartes = query.all()
        if not descartes:
            return jsonify({'error': 'Nenhum dado encontrado para exportar'}), 404
        dados = [{
            'codigo': d.codigo,
            'data_descarte': d.data_descarte.strftime('%d/%m/%Y %H:%M') if d.data_descarte else 'N/A',
            'operador': d.operador,
            'motivo': d.motivo
        } for d in descartes]

        
        wb = gerar_excel_descartes(dados)
        
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        filename = f'historico_descartes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error(f"Erro ao exportar histórico de descartes: {e}")
        return jsonify({
            'error': str(e),
            'detail': traceback.format_exc()
        }), 500 

@app.route('/api/ferramentas/importar', methods=['POST'])
@login_required
def importar_ferramentas_endpoint():
    """
    Endpoint para iniciar a importação de ferramentas do arquivo externo.
    """
    try:
        logging.info(f"[IMPORT] Usuário {current_user.nome} iniciou a importação de ferramentas.")
        
        # 1. Consumir os arquivos e extrair os dados
        resultado_consumo = consumir_ferramentas(remover_apos_processar=False)
        logging.info(f"[IMPORT] Resultado do consumo: {resultado_consumo}")

        if resultado_consumo['erros']:
            # Se houver erros na leitura, retorna a mensagem de erro
            return jsonify({'success': False, 'message': "; ".join(resultado_consumo['erros'])}), 500

        if not resultado_consumo['dados']:
            return jsonify({'success': False, 'message': 'Nenhum dado válido encontrado nos arquivos para importar.'})

        # 2. Importar os dados extraídos para o banco de dados
        adicionadas, atualizadas = importar_ferramentas_para_db(db, Ferramenta, resultado_consumo['dados'])
        logging.info(f"[IMPORT] BD: {adicionadas} adicionadas, {atualizadas} atualizadas.")

        total = adicionadas + atualizadas
        if total > 0:
            return jsonify({'success': True, 'message': f'{total} ferramentas foram importadas/atualizadas com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Nenhuma ferramenta nova ou modificada para importar.'})

    except Exception as e:
        logging.error(f"[IMPORT] Erro inesperado na rota de importação: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ocorreu um erro no servidor: {e}'}), 500
    

def formatar_numero(valor):
    try:
        numero = float(valor)  # forçar conversão, mesmo que venha como string
        return f"{numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception as e:
        print(f"Erro ao formatar número: {e}")
        return valor
@app.route('/analise')
@login_required
def analise():
    try:
        
        mes_selecionado = request.args.get('mes')
        if mes_selecionado:
            data_analise = datetime.strptime(mes_selecionado, '%Y-%m')
        else:
            data_analise = datetime.now()
        hotspots_data = {
            '1': {'top': '11.5274%', 'left': '82.969%'},
            '2': {'top': '42.4126%', 'left': '80.2335%'},
            '3': {'top': '11.2392%', 'left': '77.4408%'},
            '4': {'top': '42.1245%', 'left': '74.7099%'},
            '5': {'top': '11.8156%', 'left': '72.052%'},
            '6': {'top': '43.2772%', 'left': '69.3999%'},
            '7': {'top': '11.8156%', 'left': '66.5056%'},
            '8': {'top': '42.989%', 'left': '63.9277%'},
            '9': {'top': '10.951%', 'left': '61.1522%'},
            '10': {'top': '42.7008%', 'left': '58.58%'},
            '11': {'top': '11.8156%', 'left': '55.7315%'},
            '12': {'top': '42.7008%', 'left': '53.1558%'},
            '13': {'top': '11.5274%', 'left': '50.4237%'},
            '14': {'top': '42.4126%', 'left': '47.6483%'},
            '15': {'top': '12.1037%', 'left': '44.9756%'},
            '16': {'top': '42.7008%', 'left': '42.2971%'},
            '17': {'top': '11.5274%', 'left': '39.388%'},
            '18': {'top': '43.2772%', 'left': '36.9426%'},
            '19': {'top': '12.3919%', 'left': '34.1523%'},
            '20': {'top': '42.7008%', 'left': '31.3631%'},
            '21': {'top': '11.8156%', 'left': '28.7121%'},
            '22': {'top': '42.989%', 'left': '26.0417%'},
            '23': {'top': '12.9683%', 'left': '23.4352%'},
            '24': {'top': '43.2772%', 'left': '20.8048%'},
            '25': {'top': '12.3919%', 'left': '17.8271%'},
            '26': {'top': '43.2772%', 'left': '15.167%'},
            '27': {'top': '11.5274%', 'left': '12.4143%'},
            '28': {'top': '42.7008%', 'left': '9.7839%'}
        }

        
        inicio_mes = data_analise.replace(day=1, hour=0, minute=0, second=0)
        if data_analise.month == 12:
            fim_mes = data_analise.replace(year=data_analise.year + 1, month=1, day=1) - timedelta(seconds=1)
        else:
            fim_mes = data_analise.replace(month=data_analise.month + 1, day=1) - timedelta(seconds=1)

        
        trocas = HistoricoTroca.query.filter(
            HistoricoTroca.data.between(inicio_mes, fim_mes)
        ).all()

        descartes = DescarteFerramenta.query.filter(
            DescarteFerramenta.data_descarte.between(inicio_mes, fim_mes)
        ).all()

        manutencoes = ManutencaoFerramenta.query.filter(
            ManutencaoFerramenta.data_entrada.between(inicio_mes, fim_mes)
        ).all()
        analise_posicoes = defaultdict(lambda: {
            'total_trocas': 0,
            'total_descartes': 0,
            'total_manutencoes': 0,
            'vida_util_media': [],
            'frequencia_operacoes': 0
        })
        total_operacoes = 0
        total_posicoes_criticas = 0
        total_vida_util = []  
        mes_anterior_dados = defaultdict(lambda: {
            'total_trocas': 0,
            'total_descartes': 0,
            'total_manutencoes': 0,
            'frequencia_operacoes': 0
        })
        for troca in trocas:
            pos = str(troca.posicao)  
            analise_posicoes[pos]['total_trocas'] += 1
            if troca.producao_atual:  
                analise_posicoes[pos]['vida_util_media'].append(troca.producao_atual)
                total_vida_util.append(troca.producao_atual)
            analise_posicoes[pos]['frequencia_operacoes'] += 1
            total_operacoes += 1
        dados_analise = {}
        maior_frequencia = 0
        media_vida_util = float(np.mean(total_vida_util)) if total_vida_util else 0  

        for pos, dados in analise_posicoes.items():
            vida_util = dados['vida_util_media']
            freq = dados['frequencia_operacoes']
            maior_frequencia = max(maior_frequencia, freq)
            
            dados_analise[pos] = {
                'total_trocas': dados['total_trocas'],
                'total_descartes': dados['total_descartes'],
                'total_manutencoes': dados['total_manutencoes'],
                'vida_util_media': float(np.mean(vida_util)) if vida_util else 0,
                'frequencia_operacoes': freq
            }
        maior_frequencia = max(1, maior_frequencia)
        meses_trocas = db.session.query(
            db.func.strftime('%Y-%m', HistoricoTroca.data)
        ).distinct().all()
        meses_descartes = db.session.query(
            db.func.strftime('%Y-%m', DescarteFerramenta.data_descarte)
        ).distinct().all()
        meses_manutencoes = db.session.query(
            db.func.strftime('%Y-%m', ManutencaoFerramenta.data_entrada)
        ).distinct().all()
        todos_meses = set()
        for meses in [meses_trocas, meses_descartes, meses_manutencoes]:
            todos_meses.update([mes[0] for mes in meses if mes[0]])
        meses_disponiveis = []
        for mes in sorted(todos_meses, reverse=True):
            data = datetime.strptime(mes, '%Y-%m')
            meses_disponiveis.append({
                'valor': mes,
                'texto': data.strftime('%b/%Y')
            })
        mes_selecionado = request.args.get('mes')
        if not mes_selecionado and meses_disponiveis:
            mes_selecionado = meses_disponiveis[0]['valor']
            return redirect(url_for('analise', mes=mes_selecionado))
        total_operacoes = 0
        total_posicoes_criticas = 0
        total_vida_util = []
        mes_anterior_dados = defaultdict(lambda: {
            'total_trocas': 0,
            'total_descartes': 0,
            'total_manutencoes': 0,
            'frequencia_operacoes': 0
        })
        mes_anterior = inicio_mes - timedelta(days=1)
        inicio_mes_anterior = mes_anterior.replace(day=1, hour=0, minute=0, second=0)
        fim_mes_anterior = inicio_mes - timedelta(seconds=1)
        trocas_anterior = HistoricoTroca.query.filter(
            HistoricoTroca.data.between(inicio_mes_anterior, fim_mes_anterior)
        ).all()
        total_operacoes_anterior = 0
        for troca in trocas_anterior:
            total_operacoes_anterior += 1
            pos = str(troca.posicao)
            mes_anterior_dados[pos]['total_trocas'] += 1
            mes_anterior_dados[pos]['frequencia_operacoes'] += 1
        analise_posicoes = defaultdict(lambda: {
            'total_trocas': 0,
            'total_descartes': 0,
            'total_manutencoes': 0,
            'vida_util_media': [],
            'frequencia_operacoes': 0
        })
        for troca in trocas:
            pos = str(troca.posicao)
            analise_posicoes[pos]['total_trocas'] += 1
            if troca.producao_atual:  
                analise_posicoes[pos]['vida_util_media'].append(troca.producao_atual)  
                total_vida_util.append(troca.producao_atual)  
            analise_posicoes[pos]['frequencia_operacoes'] += 1
            total_operacoes += 1
        for descarte in descartes:
            
            ultima_troca = HistoricoTroca.query.filter_by(
                codigo=descarte.codigo
            ).order_by(HistoricoTroca.data.desc()).first()
            
            if ultima_troca:
                pos = str(ultima_troca.posicao)
                analise_posicoes[pos]['total_descartes'] += 1
                analise_posicoes[pos]['frequencia_operacoes'] += 1
                total_operacoes += 1
        for manutencao in manutencoes:
            ferramenta = Ferramenta.query.get(manutencao.ferramenta_id)
            if ferramenta:
                
                posicoes_ferramenta = HistoricoTroca.query.filter(
                    HistoricoTroca.codigo == ferramenta.codigo,
                    HistoricoTroca.data <= manutencao.data_entrada
                ).order_by(HistoricoTroca.data.desc()).first()
                
                if posicoes_ferramenta:
                    pos = str(posicoes_ferramenta.posicao)
                    analise_posicoes[pos]['total_manutencoes'] += 1
                    analise_posicoes[pos]['frequencia_operacoes'] += 1
                    total_operacoes += 1
        dados_analise = {}
        maior_frequencia = 0
        media_vida_util = float(np.mean(total_vida_util)) if total_vida_util else 0  
        for pos, dados in analise_posicoes.items():
            vida_util = dados['vida_util_media']
            freq = dados['frequencia_operacoes']
            maior_frequencia = max(maior_frequencia, freq)
            
            
            if freq > 12:
                total_posicoes_criticas += 1

            dados_analise[pos] = {
                'total_trocas': dados['total_trocas'],
                'total_descartes': dados['total_descartes'],
                'total_manutencoes': dados['total_manutencoes'],
                'vida_util_media': float(np.mean(vida_util)) if vida_util else 0,
                'frequencia_operacoes': freq
            }
        if total_operacoes_anterior > 0:
            operacoes_percentual = ((total_operacoes - total_operacoes_anterior) / total_operacoes_anterior) * 100
        else:
            operacoes_percentual = 100
        return render_template('analise.html',
                             dados_analise=dados_analise,
                             hotspots=hotspots_data,
                             total_posicoes=28,
                             maior_frequencia=maior_frequencia,
                             mes_atual=data_analise.strftime('%b/%Y'),
                             meses_disponiveis=meses_disponiveis,
                             mes_selecionado=mes_selecionado,
                             total_operacoes=total_operacoes,
                             operacoes_percentual=operacoes_percentual,
                             total_vida_util_formatado=formatar_numero(media_vida_util),
                             media_vida_util=media_vida_util,
                             total_posicoes_criticas=total_posicoes_criticas)
    
    except Exception as e:
        logging.error(f"Erro na análise: {str(e)}")
        print(f"Erro detalhado: {traceback.format_exc()}")  
        flash('Erro ao gerar análise.', 'danger')
        return redirect(url_for('index'))

@app.route('/testar_envio_email_turno', methods=['GET'])
@login_required
def testar_envio_email_turno():
    if current_user.area != 'supervisor':
        return jsonify({'error': 'Acesso não autorizado'}), 403

    turno = request.args.get('turno')
    if turno not in ['noite', 'manha']:
        return jsonify({
            'success': False, 
            'error': "Parâmetro 'turno' inválido. Use 'noite' ou 'manha'."
        }), 400

    try:
        enviar_email_relatorio_diario(turno)
        return jsonify({'success': True, 'message': f'E-mail de teste para o turno da {turno} enviado com sucesso!'}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    sao_paulo_tz = pytz.timezone('America/Sao_Paulo')
    scheduler = BackgroundScheduler(timezone=sao_paulo_tz)

    def job_noite():
        with app.app_context():
            logging.info('[SCHEDULER] Disparando envio de email do turno NOITE')
            enviar_email_relatorio_diario('noite')
            
    def job_dia():
        with app.app_context():
            logging.info('[SCHEDULER] Disparando envio de email do turno DIA')
            enviar_email_relatorio_diario('manha')

    import os
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
        scheduler.add_job(job_dia, CronTrigger(hour=7, minute=0, timezone=sao_paulo_tz), id='job_dia', replace_existing=True)
        scheduler.add_job(job_noite, CronTrigger(hour=19, minute=0, timezone=sao_paulo_tz), id='job_noite', replace_existing=True)
        scheduler.start()
        logging.info('[SCHEDULER] Jobs de email agendados.')
    else:
        logging.info('[SCHEDULER] Ignorando agendamento em processo secundário (debug/reload)')

def monitorar_pasta_ferramentas():
    from utils.ferramentas_importer import CAMINHO_FONTE, consumir_ferramentas
    logging.info(f"[MONITOR] Iniciando monitoramento da pasta: {CAMINHO_FONTE}")
    while True:
        with app.app_context():
            consumir_ferramentas(db, Ferramenta)
        time.sleep(5) 

if __name__ == '__main__':
    init_db()
    
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        monitor_thread = threading.Thread(target=monitorar_pasta_ferramentas, daemon=True)
        monitor_thread.start()

    sao_paulo_tz = pytz.timezone('America/Sao_Paulo')
    scheduler = BackgroundScheduler(timezone=sao_paulo_tz)

    def job_noite():
        with app.app_context():
            logging.info('[SCHEDULER] Disparando envio de email do turno NOITE')
            enviar_email_relatorio_diario('noite')

    def job_dia():
        with app.app_context():
            logging.info('[SCHEDULER] Disparando envio de email do turno DIA')
            enviar_email_relatorio_diario('manha')

    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
        scheduler.add_job(job_dia, CronTrigger(hour=7, minute=0, timezone=sao_paulo_tz), id='job_dia', replace_existing=True)
        scheduler.add_job(job_noite, CronTrigger(hour=19, minute=0, timezone=sao_paulo_tz), id='job_noite', replace_existing=True)
        scheduler.start()
        logging.info('[SCHEDULER] Jobs de email agendados.')
    else:
        logging.info('[SCHEDULER] Ignorando agendamento em processo secundário (debug/reload)')

    app.run(debug=False, host='0.0.0.0', port=552402)
